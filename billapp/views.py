from django.shortcuts import render
from openpyxl import load_workbook
import os
from django.conf import settings
from .forms import UserRegistrationForm
from django.contrib.auth import login
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Customer
from .forms import CustomerForm, InvoiceForm, MachineForm,ServiceNoteForm, CopyCounterForm
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.conf import settings
from openpyxl import load_workbook
import json
import os
from datetime import datetime
from decimal import Decimal
from .models import Customer, Invoice, LineItem, Machines, ServiceNote
from django.contrib.auth.decorators import login_required
import re
from datetime import datetime
from django.utils.timezone import now
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import EmailMessage
import io
import tempfile
import openpyxl
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import subprocess
import shutil
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, Avg, Count
from django.urls import reverse
from django.utils import timezone
import cloudinary.uploader


class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'customers/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 10
    ordering = ['-id']

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                name__icontains=search_query
            ) | queryset.filter(
                email__icontains=search_query
            ) | queryset.filter(
                phone__icontains=search_query
            )
        return queryset


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'customers/customer_detail.html'
    context_object_name = 'customer'


class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'customers/customer_form.html'
    success_url = reverse_lazy('customer_list')


    def form_valid(self, form):
        name = form.cleaned_data.get('name')
        email = form.cleaned_data.get('email')

        if Customer.objects.filter(name=name, email=email).exists():
            messages.error(self.request, 'Customer with this name and email already exists.')
            return self.form_invalid(form)

        messages.success(self.request, 'Customer created successfully!')
        return super().form_valid(form)


    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'customers/customer_form.html'
    success_url = reverse_lazy('customer_list')

    def form_valid(self, form):
        messages.success(self.request, 'Customer updated successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        return context


class CustomerDeleteView(LoginRequiredMixin, DeleteView):
    model = Customer
    template_name = 'customers/customer_confirm_delete.html'
    success_url = reverse_lazy('customer_list')
    context_object_name = 'customer'

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Customer deleted successfully!')
        return super().delete(request, *args, **kwargs)

def index(request):
    recent_customers = Customer.objects.order_by('-id')[:5]
    recent_invoices = Invoice.objects.all().order_by('-date')[:5]

    current_year = timezone.now().year

    
    paid_bills = Invoice.objects.filter(
        date__year=current_year,
        invoice_type=Invoice.BILL,
        payment_status=Invoice.PAID
    )
    monthly_bill_sales = (
        paid_bills.annotate(month=TruncMonth('date'))
                .values('month')
                .annotate(total=Sum('total_amount'))
                .order_by('month')
    )
    bill_sales_data = [0] * 12
    for entry in monthly_bill_sales:
        month_index = entry['month'].month - 1
        if entry['total'] is not None:
            bill_sales_data[month_index] = float(entry['total'])

    # --- 2. Calculate Monthly Service Fees ---
    service_notes = ServiceNote.objects.filter(date_of_service__year=current_year)
    monthly_service_fees = (
        service_notes.annotate(month=TruncMonth('date_of_service'))
                     .values('month')
                     .annotate(total=Sum('fee_charged'))
                     .order_by('month')
    )
    service_fee_data = [0] * 12
    for entry in monthly_service_fees:
        month_index = entry['month'].month - 1
        if entry['total'] is not None:
            service_fee_data[month_index] = float(entry['total'])

    
    combined_revenue_data = [x + y for x, y in zip(bill_sales_data, service_fee_data)]

    
    return render(request, 'index.html', {
        'recent_customers': recent_customers,
        'recent_invoices': recent_invoices,
        'monthly_revenue_data': combined_revenue_data, 
    })


    
def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password1'])
            user.save()
            login(request, user)
            return redirect('')
    else:
        form = UserRegistrationForm()

    return render(request, 'registration/register.html',{'form':form})

def generate_excel_invoice(invoice):
    """
    Generate Excel invoice and upload to Cloudinary
    """
    try:
        # Determine template name
        template_name = 'BILL_template.xlsx' if invoice.invoice_type == 'BILL' else 'QUOTATION_template.xlsx'
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'excel', template_name)
        
        # Load template
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Fill customer information
        customer_info = f"{invoice.customer.name}\n{invoice.customer.address}"
        ws['A9'] = customer_info
        
        # Fill invoice details
        ws['E8'] = invoice.date.strftime('%d/%m/%Y')
        ws['E10'] = invoice.invoice_number
        
        # Fill line items
        line_items = invoice.items.all()
        row_start = 14
        
        for i, item in enumerate(line_items):
            if i >= 6:  # Limit to 6 items as per template
                break
            
            base_row = row_start + (i * 3)
            
            # Fill item details
            ws[f'A{base_row}'] = item.product_name
            ws[f'D{base_row}'] = item.quantity
            ws[f'E{base_row}'] = float(item.rate)
            ws[f'F{base_row}'] = float(item.amount)
        
        # Fill total amount
        ws['F35'] = float(invoice.total_amount)
        
        # Generate filename
        raw_invoice_number = invoice.invoice_number
        bill_number = raw_invoice_number.split('/')[-1] if '/' in raw_invoice_number else raw_invoice_number
        customer_full_name = re.sub(r'[^\w\s]', '', invoice.customer.name).strip().replace(' ', '_')
        month_abbr = invoice.date.strftime('%b').upper()
        year = invoice.date.strftime('%Y')
        filename = f"{bill_number}_{customer_full_name}_{month_abbr}_{year}.xlsx"
        
        # Save Excel to memory buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload to Cloudinary
        if hasattr(settings, 'CLOUDINARY_STORAGE'):
            # Upload as raw file (for Excel files)
            upload_result = cloudinary.uploader.upload(
                excel_buffer.getvalue(),
                resource_type="raw",
                public_id=f"invoice-files/{filename.replace('.xlsx', '')}",
                format="xlsx",
                overwrite=True,
                use_filename=False,
            )
            
            # Return the secure URL
            return upload_result['secure_url']
        
        else:
            # Fallback to local storage for development
            bills_dir = os.path.join(settings.MEDIA_ROOT, 'bills')
            os.makedirs(bills_dir, exist_ok=True)
            file_path = os.path.join(bills_dir, filename)
            
            with open(file_path, 'wb') as f:
                f.write(excel_buffer.getvalue())
            
            return file_path
        
    except Exception as e:
        raise Exception(f"Error generating Excel file: {str(e)}")

@login_required
def invoice_create(request):
    customers = Customer.objects.all()

    
    today = timezone.now().date()
    if today.month >= 4:  
        fy_start = today.year
        fy_end = today.year + 1
    else:  
        fy_start = today.year - 1
        fy_end = today.year
    prefix = f"{str(fy_start)[-2:]}-{str(fy_end)[-2:]}"  

    
    last_invoice = Invoice.objects.filter(invoice_number__startswith=prefix).order_by('-id').first()
    if last_invoice and '/' in last_invoice.invoice_number:
        try:
            last_num = int(last_invoice.invoice_number.split('/')[-1])
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1

    suggested_invoice_number = f"{prefix}/{next_num}"

    
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            try:
                invoice = Invoice.objects.create(
                    invoice_type=form.cleaned_data['invoice_type'],
                    invoice_number=form.cleaned_data['invoice_number'],
                    customer=form.cleaned_data['customer'],
                    date=form.cleaned_data['date']
                )
                
                total_amount = Decimal('0.00')
                for i in range(1, 7): 
                    product_name = form.cleaned_data.get(f'product_name_{i}')
                    quantity = form.cleaned_data.get(f'quantity_{i}')
                    rate = form.cleaned_data.get(f'rate_{i}')
                    
                    if product_name and quantity and rate:
                        amount = quantity * rate
                        LineItem.objects.create(
                            invoice=invoice,
                            product_name=product_name,
                            quantity=quantity,
                            rate=rate,
                            amount=amount
                        )
                        total_amount += amount
                
                invoice.total_amount = total_amount
                invoice.save()
                
                file_path = generate_excel_invoice(invoice)
                invoice.file_path = file_path
                invoice.save()
                
                messages.success(request, f'{invoice.invoice_type} created successfully!')
                return redirect('invoice_detail', pk=invoice.pk)
            except Exception as e:
                messages.error(request, f'Error creating invoice: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        
        form = InvoiceForm(initial={'invoice_number': suggested_invoice_number})

    return render(request, 'invoices/invoice_form.html', {
        'form': form,
        'customers': customers
    })

@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    return render(request, 'invoices/invoice_detail.html', {'invoice': invoice})

@login_required
def invoice_list(request):
    
    invoices = Invoice.objects.all().order_by('-date')
    
    
    search_query = request.GET.get('search', '')
    invoice_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    amount_min = request.GET.get('amount_min', '')
    amount_max = request.GET.get('amount_max', '')
    mailed_status = request.GET.get('mailed_status', '')
    
    
    if search_query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search_query) |
            Q(customer__name__icontains=search_query) |
            Q(customer__email__icontains=search_query)
        )
    
    if invoice_type:
        invoices = invoices.filter(invoice_type=invoice_type)
    
    if date_from:
        try:
            date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
            invoices = invoices.filter(date__gte=date_from_parsed)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
            invoices = invoices.filter(date__lte=date_to_parsed)
        except ValueError:
            pass
    
    if amount_min:
        try:
            amount_min_decimal = Decimal(amount_min)
            invoices = invoices.filter(total_amount__gte=amount_min_decimal)
        except (ValueError, TypeError):
            pass
    
    if amount_max:
        try:
            amount_max_decimal = Decimal(amount_max)
            invoices = invoices.filter(total_amount__lte=amount_max_decimal)
        except (ValueError, TypeError):
            pass
    
    if mailed_status:
        if mailed_status == 'true':
            invoices = invoices.filter(mailed=True)
        elif mailed_status == 'false':
            invoices = invoices.filter(mailed=False)
    
    
    paginator = Paginator(invoices, 10)  
    page = request.GET.get('page')
    
    try:
        invoices = paginator.page(page)
    except PageNotAnInteger:
        
        invoices = paginator.page(1)
    except EmptyPage:
        
        invoices = paginator.page(paginator.num_pages)
    
    context = {
        'invoices': invoices,
        'search_query': search_query,
        'invoice_type': invoice_type,
        'date_from': date_from,
        'date_to': date_to,
        'amount_min': amount_min,
        'amount_max': amount_max,
        'mailed_status': mailed_status,
        'is_paginated': invoices.has_other_pages(),
        'page_obj': invoices,
    }
    
    return render(request, 'invoices/invoice_list.html', context)

def mark_invoice_mailed(request, pk):
    if request.method == 'POST':
        try:
            invoice = get_object_or_404(Invoice, pk=pk)
            invoice.mailed = True
            invoice.save()
            return JsonResponse({
                'success': True,
                'message': f'{invoice.get_invoice_type_display()} marked as mailed successfully!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

def download_invoice(request, pk):
    """
    Updated download function for Cloudinary-stored files
    """
    from django.shortcuts import get_object_or_404, redirect
    from django.contrib import messages
    from django.http import HttpResponse
    import requests
    
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if invoice.file_path:
        try:
            if invoice.file_path.startswith('http'):
                # It's a Cloudinary URL, download and serve
                response = requests.get(invoice.file_path)
                if response.status_code == 200:
                    filename = f"{invoice.invoice_number.replace('/', '_')}.xlsx"
                    http_response = HttpResponse(
                        response.content,
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    http_response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    return http_response
                else:
                    messages.error(request, 'File not found on server.')
            else:
                # Local file (development)
                if os.path.exists(invoice.file_path):
                    with open(invoice.file_path, 'rb') as f:
                        response = HttpResponse(
                            f.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(invoice.file_path)}"'
                        return response
                else:
                    messages.error(request, 'File not found.')
        except Exception as e:
            messages.error(request, f'Error downloading file: {str(e)}')
    else:
        messages.error(request, 'No file associated with this invoice.')
    
    return redirect('invoice_detail', pk=pk)
    
class InvoiceDeleteView(DeleteView):
    model = Invoice
    template_name = "invoices/invoice_confirm_delete.html"
    success_url = reverse_lazy("invoice_list")  

    
def libreoffice_convert_to_pdf(excel_file):
    soffice_path = "soffice"

    if not os.path.exists(soffice_path):
        raise FileNotFoundError(f"LibreOffice not found at {soffice_path}")

    output_dir = tempfile.mkdtemp()

    subprocess.run([
        soffice_path, "--headless", "--convert-to", "pdf",
        "--outdir", output_dir, excel_file
    ], check=True)

    filename = os.path.splitext(os.path.basename(excel_file))[0] + ".pdf"
    pdf_path = os.path.join(output_dir, filename)

    if not os.path.exists(pdf_path):
        raise FileNotFoundError("PDF not generated by LibreOffice.")

    return pdf_path

def send_invoice_email(invoice):

    try:
        excel_file = invoice.file_path
        pdf_file = libreoffice_convert_to_pdf(excel_file)

        with open(pdf_file, "rb") as f:
            pdf_content = f.read()

        subject = f"Invoice #{invoice.invoice_number}"
        body = (
            f"Dear {invoice.customer.name},\n\n"
            f"Please find attached your invoice #{invoice.invoice_number}.\n\n"
            f"Thank you for your business.\n\n"
            f"Best regards,\nBilling Team"
        )
        email = EmailMessage(
            subject,
            body,
            settings.DEFAULT_FROM_EMAIL,
            [invoice.customer.email],
        )
        email.attach(
            f"Invoice_{invoice.invoice_number}.pdf",
            pdf_content,
            "application/pdf"
        )
        email.send()

        invoice.mailed = True
        invoice.save()
        return True

    except Exception as e:
        print(f"Error sending invoice email: {e}")
        return False


@login_required
def send_invoice_email_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    success = send_invoice_email(invoice)

    if success:
        return JsonResponse({"status": "success", "message": f"Invoice #{invoice.invoice_number} email sent successfully!"})
    else:
        return JsonResponse({"status": "error", "message": f"Failed to send email for Invoice #{invoice.invoice_number}."})
    

@csrf_exempt
def update_payment_status(request, pk):
    if request.method == "POST":
        invoice = get_object_or_404(Invoice, pk=pk)
        if invoice.payment_status != 'PAID':  # Fixed condition
            invoice.payment_status = 'PAID'
            invoice.save()
            return JsonResponse({"success": True, "message": "Payment marked as Paid"})
        else:
            return JsonResponse({"success": False, "message": "Invoice already marked as Paid"})
    return JsonResponse({"success": False, "message": "Invalid request"})


def add_machine(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)
    
    if request.method == 'POST':
        form = MachineForm(request.POST)
        if form.is_valid():
            
            machine = form.save(commit=False)
            
           
            machine.customer = customer
            
            
            purchase_date = form.cleaned_data['purchase_date']
            machine.warranty_expiry = purchase_date + relativedelta(months=+6)
            
            
            machine.save()
            
            
            return redirect('customer_detail', pk=customer.id)
    else:
        form = MachineForm()
        
    context = {
        'form': form,
        'customer': customer
    }
    return render(request, 'service_tracking/machine_form.html', context)

def machine_update(request, pk):
    machine = get_object_or_404(Machines, pk=pk)
    if request.method == 'POST':
        form = MachineForm(request.POST, instance=machine)
        if form.is_valid():
            
            updated_machine = form.save(commit=False)
            if 'purchase_date' in form.changed_data:
                purchase_date = form.cleaned_data['purchase_date']
                updated_machine.warranty_expiry = purchase_date + relativedelta(months=+6)
            updated_machine.save()
            return redirect('customer_detail', pk=machine.customer.pk)
    else:
        form = MachineForm(instance=machine)
        
    context = {
        'form': form,
        'is_update': True, 
        'machine': machine
    }
    return render(request, 'service_tracking/machine_form.html', context)


def machine_delete(request, pk):
    machine = get_object_or_404(Machines, pk=pk)
    customer_pk = machine.customer.pk  
    
    if request.method == 'POST':
        machine.delete()
        return redirect('customer_detail', pk=customer_pk)
        
    context = {
        'machine': machine
    }
    return render(request, 'service_tracking/machine_confirm_delete.html', context)

def machine_detail(request, pk):
    """Display detailed view of a machine with all its information and service notes"""
    machine = get_object_or_404(Machines, pk=pk)
    
    context = {
        'machine': machine,
    }
    return render(request, 'service_tracking/machine_detail.html', context)

def select_machine_for_note(request):
    if request.method == 'POST':
        machine_id = request.POST.get('machine_id')
        next_url = request.POST.get('next', '')
        if machine_id:
            add_note_url = reverse('add_service_note', args=[machine_id])
            if next_url:
                add_note_url += f"?next={next_url}"
            return redirect(add_note_url)

    customers = Customer.objects.all()
    return render(request, "service_tracking/select_machine_for_note.html", {"customers": customers})

def get_machines(request):
    customer_id = request.GET.get("customer_id")
    machines = Machines.objects.filter(customer_id=customer_id)
    data = [{"id": m.id, "name": m.machine_name, "serial_number": m.serial_number} for m in machines]
    return JsonResponse(data, safe=False)


def service_notes_view(request):
    """
    View to display all service notes with machine count and fee information
    """
    
    service_notes = ServiceNote.objects.select_related(
        'machine', 
        'machine__customer'
    ).order_by('-created_at')
    
    
    stats = service_notes.aggregate(
        total_notes=Count('id'),
        total_revenue=Sum('fee_charged'),
        average_fee=Avg('fee_charged')
    )
    
    context = {
        'service_notes': service_notes,
        'total_service_notes': stats['total_notes'] or 0,
        'total_service_revenue': stats['total_revenue'] or 0,
        'average_service_fee': stats['average_fee'] or 0,
    }
    
    return render(request, 'service_tracking/service_notes.html', context)

def add_service_note(request, machine_pk):
    """Add a new service note for a machine"""
    machine = get_object_or_404(Machines, pk=machine_pk)
    
    if request.method == 'POST':
        form = ServiceNoteForm(request.POST)
        if form.is_valid():
            service_note = form.save(commit=False)
            service_note.machine = machine
            service_note.copy_counter_at_service = machine.copy_counter
            service_note.save()
            messages.success(request, 'Service note added successfully!')

           
            next_url = request.POST.get('next')
            if next_url:
                return redirect(next_url)
            referer = request.META.get('HTTP_REFERER')
            if referer:
                return redirect(referer)
            return redirect('machine_detail', pk=machine.pk)
    else:
        form = ServiceNoteForm()
    
    context = {
        'form': form,
        'machine': machine,
        'is_update': False,
        'next': request.GET.get('next', ''),
    }
    return render(request, 'service_tracking/service_note_form.html', context)

def update_service_note(request, pk):
    """Update an existing service note"""
    service_note = get_object_or_404(ServiceNote, pk=pk)
    machine = service_note.machine
    
    if request.method == 'POST':
        form = ServiceNoteForm(request.POST, instance=service_note)
        if form.is_valid():
            form.save()
            messages.success(request, 'Service note updated successfully!')

            # Redirect dynamically
            next_url = request.POST.get('next')
            if next_url:
                return redirect(next_url)
            referer = request.META.get('HTTP_REFERER')
            if referer:
                return redirect(referer)
            return redirect('machine_detail', pk=machine.pk)
    else:
        form = ServiceNoteForm(instance=service_note)
    
    context = {
        'form': form,
        'machine': machine,
        'service_note': service_note,
        'is_update': True,
        'next': request.GET.get('next', ''),
    }
    return render(request, 'service_tracking/service_note_form.html', context)

def delete_service_note(request, pk):
    """Delete a service note with confirmation"""
    service_note = get_object_or_404(ServiceNote, pk=pk)
    machine = service_note.machine

    if request.method == 'POST':
        service_note.delete()
        messages.success(request, 'Service note deleted successfully!')

        # First try explicit "next"
        next_url = request.POST.get('next')
        if next_url:
            return redirect(next_url)

        # Fallbacks
        referer = request.META.get('HTTP_REFERER')
        if referer:
            return redirect(referer)

        return redirect('machine_detail', pk=machine.pk)

    context = {
        'service_note': service_note,
        'machine': machine
    }
    return render(request, 'service_tracking/service_note_confirm_delete.html', context)

def update_copy_counter(request, machine_pk):
    """Update the copy counter for a machine"""
    machine = get_object_or_404(Machines, pk=machine_pk)
    
    if request.method == 'POST':
        form = CopyCounterForm(request.POST, instance=machine)
        if form.is_valid():
            form.save()
            messages.success(request, f'Copy counter updated to {machine.copy_counter}!')
            return redirect('machine_detail', pk=machine.pk)
    else:
        form = CopyCounterForm(instance=machine)
    
    context = {
        'form': form,
        'machine': machine
    }
    return render(request, 'service_tracking/copy_counter_form.html', context)

def reset_copy_counter(request, machine_pk):
    """Reset the copy counter to zero with confirmation"""
    machine = get_object_or_404(Machines, pk=machine_pk)
    
    if request.method == 'POST':
        machine.copy_counter = 0
        machine.save()
        messages.success(request, 'Copy counter has been reset to 0!')
        return redirect('machine_detail', pk=machine.pk)
    
    context = {
        'machine': machine
    }
    return render(request, 'service_tracking/copy_counter_reset_confirm.html', context)